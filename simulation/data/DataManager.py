from calculations.pixels_calculations import PixelsConverter
from datetime import datetime
import threading
import pandas as pd
from openpyxl import load_workbook
pd.set_option('display.precision', 2)


class DataManager:
    def __init__(self, filename='simulation_data.xlsx', export_interval=5, roadType='', numOfLanes=1):
        self.filename = filename
        self.export_interval = export_interval
        self.roadType = roadType
        self.numOfLanes = numOfLanes
        self.statsData = pd.DataFrame(columns=['Time Stamp', 'Average Speed (km\h)', 'Density'])
        self.accidentsData = pd.DataFrame(columns=['Time Stamp', 'Vehicle Types', 'Speeds At Crash Time (km\h)', 'Accident ID'])
    
    def update_stats(self, vehicles, roadType : str, finalCall=False):
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        totalSpeed, count, totalVehiclesInJunction, totalVehiclesInRoundabout = 0, 0, 0, 0
        timeWaitedToEnterJunction = []
        roadDirectionStats = {}

        for vehicle in vehicles:
            if not vehicle.inAccident:
                speed = PixelsConverter.convert_pixels_per_frames_to_speed(vehicle.speed)
                totalSpeed += speed
                count += 1
                if vehicle.countForJunctionData:
                    totalVehiclesInJunction += 1
                    vehicle.countForJunctionData = False
                if vehicle.countForRoundaboutData:
                    totalVehiclesInRoundabout += 1
                    vehicle.countForRoundaboutData = False
                if vehicle.countForJunctionTimeData:
                    timeWaitedToEnterJunction.append(vehicle.timeWaitedToEnterJunction)
                    vehicle.countForJunctionTimeData = False
                
                key = (vehicle.roadIndex + 1, vehicle.directionIndex)
                if key not in roadDirectionStats:
                    roadDirectionStats[key] = {'total_speed': 0, 'count': 0}

                roadDirectionStats[key]['total_speed'] += speed
                roadDirectionStats[key]['count'] += 1

        amountOfVehiclesInJunctionInGivenTime = {"Vehicles passing through plus-junction": [totalVehiclesInJunction]} if roadType == "junction" else {}
        avgTimeWaitedToEnterJunction = {"Avg. time (seconds) waited to enter junction": [sum(timeWaitedToEnterJunction) / len(timeWaitedToEnterJunction)] if len(timeWaitedToEnterJunction) != 0 else "-"}
        avgTimeWaitedToEnterJunction = avgTimeWaitedToEnterJunction if roadType == "junction" else {}
        amountOfVehiclesInRoundaboutInGivenTime = {"Vehicles passing through roundabout": [totalVehiclesInRoundabout]} if roadType == "roundabout" else {}
        avgSpeed = totalSpeed / count if count > 0 else 0
        avgSpeedsPerRoadDirection = {
            f'Avg. Speed (Road {road}, Direction {direction})': (stats['total_speed'] / stats['count'] if stats['count'] > 0 else "-")
            for (road, direction), stats in roadDirectionStats.items()
        }

        new_data = pd.DataFrame({
            'Time Stamp': [current_time],
            'Average Speed (km\h)': [avgSpeed],
            'Density': [count],
            **amountOfVehiclesInJunctionInGivenTime,
            **avgTimeWaitedToEnterJunction,
            **amountOfVehiclesInRoundaboutInGivenTime,
            **avgSpeedsPerRoadDirection 
        })

        self.statsData = pd.concat([self.statsData, new_data], ignore_index=True)
        if finalCall:
            self.export_to_excel(self.roadType, self.numOfLanes)
        else:
            threading.Thread(self.export_to_excel(self.roadType, self.numOfLanes))
        
    
    def log_accident(self, vehicle_1_Type : str, vehicle_2_Type : str, vehicle_1_Speed : float, vehicle_2_Speed : float, accidentID : int):
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        vehicle_types = f"{vehicle_1_Type} and {vehicle_2_Type}"
        speeds = "{:.2f} and {:.2f}".format(vehicle_1_Speed, vehicle_2_Speed)
        new_data = pd.DataFrame({
            'Time Stamp': [current_time],
            'Vehicle Types': [vehicle_types],
            'Speeds At Crash Time (km\h)': [speeds],
            'Accident ID': [accidentID]
        })
        self.accidentsData = pd.concat([self.accidentsData, new_data], ignore_index=True)

    
    def export_to_excel(self, roadType: str, numOfLanes: int):
        sheet_name = f"{roadType} road {numOfLanes} lanes" if numOfLanes > 1 else f"{roadType} road {numOfLanes} lane"

        try:
            book = load_workbook(self.filename)
            if sheet_name in book.sheetnames:
                del book[sheet_name]
        except FileNotFoundError:
            book = None

        with pd.ExcelWriter(self.filename, engine='openpyxl', mode='a' if book else 'w') as writer:
            if book:
                writer._book = book 
            self.statsData.to_excel(writer, sheet_name=sheet_name, index=False)
            if 'Accidents' not in writer.book.sheetnames:
                self.accidentsData.to_excel(writer, sheet_name='Accidents', index=False)

        if book is not None:
            book.save(self.filename)
            book.close()